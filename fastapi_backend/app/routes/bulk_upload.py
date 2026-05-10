"""
Bulk-upload verification items from an Excel (.xlsx) file.

Expected columns (case-insensitive, extra columns ignored):
  title*          – item description (required)
  item_type       – vehicle | property | equipment | inventory | bank_account | financial_record | other
  reference_value – book value / amount claimed in client records
  description     – additional notes

Download a template from GET /audits/{id}/bulk-upload-template
"""
import io
from typing import Any

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.models.audit import Audit
from app.models.user import User
from app.models.verification_item import ItemType, VerificationItem, VerificationStatus
from app.routes.auth import get_current_user
from app.utils.audit_log import log_action

router = APIRouter(tags=["bulk_upload"])

_VALID_TYPES = {t.value for t in ItemType}


def _normalise(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


@router.get("/audits/{audit_id}/bulk-upload-template")
def download_template(
    audit_id: int,
    current_user: User = Depends(get_current_user),
):
    """Return a blank Excel template for bulk item upload."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Verification Items"
    headers = ["title", "item_type", "reference_value", "description"]
    ws.append(headers)
    ws.append(["Example: Office Vehicle KA-01-AB-1234", "vehicle", "8,50,000", "Company car purchased FY23"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bulk_items_template.xlsx"},
    )


@router.post("/audits/{audit_id}/bulk-upload-items")
async def bulk_upload_items(
    audit_id: int,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Create multiple VerificationItems from an Excel file."""
    audit = db.query(Audit).filter(
        Audit.id == audit_id,
        Audit.workspace_id == current_user.workspace_id,
    ).first()
    if not audit:
        raise HTTPException(status_code=404, detail="Audit not found")

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not parse the Excel file. Ensure it is a valid .xlsx.")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Sheet has no data rows (only a header or is empty).")

    # Map header names → column indices
    raw_headers = [_normalise(h).lower() for h in (rows[0] or [])]
    col = {name: raw_headers.index(name) for name in ("title",) if name in raw_headers}
    if "title" not in col:
        raise HTTPException(status_code=400, detail="Excel must have a 'title' column header.")

    opt_col = {}
    for name in ("item_type", "reference_value", "description"):
        if name in raw_headers:
            opt_col[name] = raw_headers.index(name)

    created = []
    failed = []
    for row_idx, row in enumerate(rows[1:], start=2):
        title = _normalise(row[col["title"]] if len(row) > col["title"] else "")
        if not title:
            failed.append({"row": row_idx, "reason": "Empty title"})
            continue

        raw_type = _normalise(row[opt_col["item_type"]] if "item_type" in opt_col and len(row) > opt_col["item_type"] else "").lower()
        item_type = ItemType(raw_type) if raw_type in _VALID_TYPES else ItemType.OTHER

        ref_val = _normalise(row[opt_col["reference_value"]] if "reference_value" in opt_col and len(row) > opt_col["reference_value"] else "") or None
        desc = _normalise(row[opt_col["description"]] if "description" in opt_col and len(row) > opt_col["description"] else "") or None

        item = VerificationItem(
            audit_id=audit_id,
            title=title[:255],
            description=desc,
            item_type=item_type,
            reference_value=ref_val,
            status=VerificationStatus.PENDING,
            workspace_id=audit.workspace_id,
        )
        db.add(item)
        created.append(item)

    if created:
        db.flush()
        log_action(
            db,
            action="imported",
            entity_type="audit",
            entity_id=audit_id,
            audit_id=audit_id,
            description=f"Bulk Excel upload: {len(created)} items created, {len(failed)} rows skipped",
            user_id=current_user.id,
            user_name=current_user.name,
            workspace_id=audit.workspace_id,
        )

    db.commit()
    for item in created:
        db.refresh(item)

    return {
        "created": len(created),
        "failed": len(failed),
        "failed_rows": failed,
        "items": [
            {"id": i.id, "title": i.title, "item_type": i.item_type, "status": i.status}
            for i in created
        ],
    }
